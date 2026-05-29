"""
来料检验记录系统（跳批检验逻辑）
IQC Incoming Inspection Record System with Skip-Lot Logic
部署：GitHub + Streamlit Cloud ; 数据存储：Google Sheets
"""

import streamlit as st
import pandas as pd
import json
import io
from datetime import datetime, date
import gspread
from google.oauth2.service_account import Credentials

st.set_page_config(page_title="来料检验记录系统", page_icon="📋", layout="wide")

# ---------- 自定义样式（深蓝表头，接近Excel） ----------
st.markdown("""
<style>
.block-container { padding-top: 2rem; max-width: 1400px; }
.tbl-header {
    background-color: #1F4E79; color: white; font-weight: bold;
    text-align: center; padding: 8px 4px; border: 1px solid #FFFFFF;
    font-size: 13px; line-height: 1.3; border-radius: 2px;
}
.tbl-cell {
    text-align: center; padding: 8px 4px; border: 1px solid #D0D0D0;
    font-size: 13px; background-color: #F8FBFF; min-height: 38px; line-height: 1.4;
}
.action-normal { color: #C0392B; font-weight: bold; }
.action-skip   { color: #1E8449; font-weight: bold; }
.action-skip-inspect { color: #1E8449; font-weight: bold; }
div[data-testid="stHorizontalBlock"] { gap: 0.4rem; }
</style>
""", unsafe_allow_html=True)

# 8个可能的测量字段（按料号动态显示，存表时固定为列）
MEASURE_COLS = ["仿形间隙测量数据","轮毂外径尺寸","内径尺寸","裙边厚度",
                "裙边高度","轴套总长","轮毂间隙","扇叶重量"]

COLUMNS = ["到货日期","订单号","供应商","零件料号","零件名称","生产日期",
           "来料总数量","检验数量","累计批次数","执行动作",
           "开始时间","结束时间","检验用时(分钟)","结果","不良备注"] \
          + MEASURE_COLS + ["检验员","记录时间"]
INSPECTORS = ["杨明","田志高","其他"]
RESULTS = ["OK","NG"]
PO_PREFIXES = ["ST-PO", "SZ-PO"]
CONSECUTIVE_PASS_TO_START = 3
SKIP_PATTERN_SKIP = 2
SKIP_PATTERN_INSPECT = 1

# 参与跳批检验的供应商白名单（其他供应商每批正常检验，不显示执行动作）
SKIP_LOT_SUPPLIERS = {"速锐达(SP)", "AQ", "金*"}

@st.cache_data
def load_parts_data():
    with open("parts_data.json","r",encoding="utf-8") as f:
        return json.load(f)
PARTS_DATA = load_parts_data()
SUPPLIERS = list(PARTS_DATA.keys())

@st.cache_resource
def get_gsheet():
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
    client = gspread.authorize(creds)
    sh = client.open_by_key(st.secrets["sheet"]["key"])
    ws = sh.sheet1
    values = ws.get_all_values()
    if not values:
        # 空表，写入表头
        ws.append_row(COLUMNS)
    elif values[0] != COLUMNS:
        # 第一行不是当前正确表头：直接用正确表头覆盖第一行
        # （注意：不insert，而是update第1行，避免插出多余的旧表头行）
        ws.update([COLUMNS], "A1", value_input_option="USER_ENTERED")
    return ws

def load_history(ws):
    """读取数据。强制用 COLUMNS 作为列名，按位置对齐，避免表头错位/重复导致读不出。"""
    values = ws.get_all_values()
    if not values or len(values) < 2:
        return pd.DataFrame(columns=COLUMNS)
    rows = values[1:]  # 跳过表头行
    ncol = len(COLUMNS)
    # 补齐/截断每行到 COLUMNS 长度
    norm_rows = []
    for r in rows:
        if len(r) < ncol:
            r = r + [""] * (ncol - len(r))
        elif len(r) > ncol:
            r = r[:ncol]
        norm_rows.append(r)
    df = pd.DataFrame(norm_rows, columns=COLUMNS)
    # 去掉完全空白的行
    df = df[df.apply(lambda x: any(str(v).strip() for v in x), axis=1)].reset_index(drop=True)
    return df

def append_record(ws, row_dict):
    ws.append_row([str(row_dict.get(c,"")) for c in COLUMNS])

def overwrite_all(ws, df):
    """用df完整覆盖整个工作表（含表头）。用于修改/删除后回写。"""
    ws.clear()
    data = [COLUMNS] + df[COLUMNS].astype(str).values.tolist()
    ws.update(data, value_input_option="USER_ENTERED")

def compute_action_and_cumulative(history_df, supplier, part_number, production_date):
    # 计算累计批次数（所有供应商都算）
    if history_df.empty:
        cumulative = 1
    else:
        mask = ((history_df["供应商"].astype(str)==str(supplier)) &
                (history_df["零件料号"].astype(str)==str(part_number)) &
                (history_df["生产日期"].astype(str)==str(production_date)))
        seq = history_df[mask].copy()
        cumulative = len(seq) + 1

    # 非白名单供应商：不参加跳批，执行动作返回空
    if supplier not in SKIP_LOT_SUPPLIERS:
        return cumulative, ""

    # 白名单供应商：按跳批状态机判定
    if history_df.empty or cumulative == 1:
        return cumulative, "正常检验"

    mask = ((history_df["供应商"].astype(str)==str(supplier)) &
            (history_df["零件料号"].astype(str)==str(part_number)) &
            (history_df["生产日期"].astype(str)==str(production_date)))
    seq = history_df[mask].copy()
    state="正常检验"; cp=0; sk=0
    for _, rec in seq.iterrows():
        is_pass = str(rec.get("结果","")).strip().upper()=="OK"
        if state=="正常检验":
            if is_pass:
                cp += 1
                if cp >= CONSECUTIVE_PASS_TO_START:
                    state="跳批"; sk=0
            else:
                cp = 0
        else:
            if is_pass:
                sk = (sk+1) % (SKIP_PATTERN_SKIP + SKIP_PATTERN_INSPECT)
            else:
                state="正常检验"; cp=0; sk=0
    if state=="正常检验":
        action="正常检验"
    else:
        action = "跳批（不检验尺寸）" if sk < SKIP_PATTERN_SKIP else "跳批检验（全项目）"
    return cumulative, action


def recalculate_all(df):
    """
    对整个历史记录重新计算"累计批次数"和"执行动作"。
    按 (供应商, 料号, 生产日期) 分组，组内按记录时间（或原顺序）排序后
    用跳批状态机重新推导每一批的累计批次和动作。
    用于删除记录后修复后续批次的累计逻辑。
    """
    if df.empty:
        return df
    df = df.copy().reset_index(drop=True)
    # 保持原始顺序作为序列顺序（记录时间升序更准，若无则按现有顺序）
    if "记录时间" in df.columns and df["记录时间"].astype(str).str.strip().ne("").any():
        df["_ord"] = pd.to_datetime(df["记录时间"], errors="coerce")
        df = df.sort_values("_ord", kind="stable").reset_index(drop=True)

    new_cum = [0] * len(df)
    new_act = [""] * len(df)

    # 分组key
    keys = list(zip(df["供应商"].astype(str), df["零件料号"].astype(str), df["生产日期"].astype(str)))
    # 每个组维护一个状态
    group_state = {}  # key -> dict(state, cp, sk, count)

    for i in range(len(df)):
        k = keys[i]
        supplier_i = k[0]
        if k not in group_state:
            group_state[k] = {"state": "正常检验", "cp": 0, "sk": 0, "count": 0}
        gs = group_state[k]

        # 当前这批的累计批次 = 组内已出现数 + 1（所有供应商都算）
        gs["count"] += 1
        new_cum[i] = gs["count"]

        # 非白名单供应商：执行动作留空，跳过跳批状态推进
        if supplier_i not in SKIP_LOT_SUPPLIERS:
            new_act[i] = ""
            continue

        # 白名单供应商：按跳批状态机判定执行动作
        if gs["state"] == "正常检验":
            new_act[i] = "正常检验"
        else:
            new_act[i] = "跳批（不检验尺寸）" if gs["sk"] < SKIP_PATTERN_SKIP else "跳批检验（全项目）"

        # 根据本批结果推进状态（供下一批用）
        is_pass = str(df.iloc[i].get("结果", "")).strip().upper() == "OK"
        if gs["state"] == "正常检验":
            if is_pass:
                gs["cp"] += 1
                if gs["cp"] >= CONSECUTIVE_PASS_TO_START:
                    gs["state"] = "跳批"
                    gs["sk"] = 0
            else:
                gs["cp"] = 0
        else:
            if is_pass:
                gs["sk"] = (gs["sk"] + 1) % (SKIP_PATTERN_SKIP + SKIP_PATTERN_INSPECT)
            else:
                gs["state"] = "正常检验"
                gs["cp"] = 0
                gs["sk"] = 0

    df["累计批次数"] = new_cum
    df["执行动作"] = new_act
    if "_ord" in df.columns:
        df = df.drop(columns=["_ord"])
    return df


def get_production_dates(history_df, supplier, part_number):
    if history_df.empty:
        return []
    mask = ((history_df["供应商"].astype(str)==str(supplier)) &
            (history_df["零件料号"].astype(str)==str(part_number)))
    dates = history_df[mask]["生产日期"].astype(str).unique().tolist()
    return sorted([d for d in dates if d and d!="nan"], reverse=True)


def make_excel(df):
    """把DataFrame导出为带格式的Excel（深蓝表头，接近Excel原表样式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "来料检验记录"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cols = list(df.columns)
    # 表头
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 26

    # 数据
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=ri, column=ci, value=str(row[col]))
            c.font = cell_font
            c.alignment = center
            c.border = border

    # 列宽自适应
    for ci, col in enumerate(cols, 1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 1.6, 10), 40)

    # 冻结表头
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ---------- 界面 ----------
st.title("📋 来料检验记录系统")
st.markdown(
    '<p style="color:#C0392B; font-size:14px; margin-top:-10px;">'
    '※ 只有 AQ、SP、金* 参加跳批检验，其他供应商每批正常检验</p>',
    unsafe_allow_html=True,
)

try:
    ws = get_gsheet()
    history_df = load_history(ws)
    gsheet_ok = True
except Exception as e:
    st.error(f"⚠️ Google Sheets 连接失败：{e}")
    st.info("请检查 Streamlit secrets 中的 gcp_service_account 和 sheet.key 配置。")
    history_df = pd.DataFrame(columns=COLUMNS)
    gsheet_ok = False

tab1, tab2 = st.tabs(["✍️ 新增检验记录", "📜 历史记录"])

with tab1:
    st.markdown("##### 录入数据")

    # 第一排：到货日期、订单号、供应商、零件料号
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        arrival_date = st.date_input("到货日期", value=date.today())
    with c2:
        pc1, pc2 = st.columns([1, 1.4])
        with pc1:
            po_prefix = st.selectbox("订单前缀", PO_PREFIXES, label_visibility="visible")
        with pc2:
            po_number = st.text_input("订单号(数字)", "", placeholder="如 12345")
        order_no = f"{po_prefix}-{po_number.strip()}" if po_number.strip() else ""
    with c3:
        supplier = st.selectbox("供应商", SUPPLIERS, index=0)
    with c4:
        parts_list = PARTS_DATA.get(supplier, [])
        part_options, part_display_map = [], {}
        for p in parts_list:
            pn = p["part_number"]; pname = p.get("part_name","")
            disp = f"{pn} | {pname}" if pname else pn
            part_options.append(disp); part_display_map[disp] = p
        if part_options:
            part_disp = st.selectbox("零件料号", part_options, index=0)
            selected_part = part_display_map[part_disp]
            part_number = selected_part["part_number"]
            part_name = selected_part.get("part_name","")
            measure_fields = selected_part.get("measure_fields", [])
        else:
            st.warning("该供应商暂无料号"); part_number=""; part_name=""; measure_fields=[]

    # 第二排：生产日期、来料总数量、检验数量、检验员
    c5, c6, c7, c8 = st.columns(4)
    with c5:
        production_date = str(st.date_input("生产日期", value=date.today(), key="prod_cal"))
    with c6:
        total_qty = st.number_input("来料总数量", min_value=0, value=0, step=1)
    with c7:
        inspect_qty = st.number_input("检验数量", min_value=0, value=0, step=1)
    with c8:
        inspector_sel = st.selectbox("检验员", INSPECTORS)
        if inspector_sel=="其他":
            inspector = st.text_input("输入姓名", "", label_visibility="collapsed", placeholder="请输入检验员姓名")
        else:
            inspector = inspector_sel

    # 第三排：开始时间、结束时间、结果
    c9, c10, c11, c12 = st.columns(4)
    with c9:
        start_time = st.time_input("开始时间", value=datetime.now().time())
    with c10:
        end_time = st.time_input("结束时间", value=datetime.now().time())
    with c11:
        result = st.selectbox("结果", RESULTS)
    with c12:
        st.empty()

    # 动态测量字段（根据料号显示）
    measure_values = {}
    if measure_fields:
        st.markdown("##### 📐 测量数据（根据所选料号显示）")
        mcols = st.columns(min(len(measure_fields), 4))
        for idx, field in enumerate(measure_fields):
            with mcols[idx % len(mcols)]:
                measure_values[field] = st.text_input(field, "", placeholder="填入数值，如 12.5")

    # NG 时弹出不良内容备注框
    if result == "NG":
        defect_note = st.text_area(
            "⚠️ 不良内容（NG必填）",
            value="",
            placeholder="请填写不良现象、缺陷位置、数量等，例如：表面划伤2处 / 装配孔位偏移0.5mm / 镀层局部脱落",
            height=80,
        )
    else:
        defect_note = ""

    cumulative, action = compute_action_and_cumulative(history_df, supplier, part_number, production_date)
    today = date.today()
    dt_start = datetime.combine(today, start_time)
    dt_end = datetime.combine(today, end_time)
    diff_min = (dt_end - dt_start).total_seconds()/60
    if diff_min < 0: diff_min += 24*60

    if not action:
        # 非白名单供应商：执行动作为空
        action_html = '<span style="color:#999;">—</span>'
    elif "跳批（不检验尺寸）" in action:
        action_html = f'<span class="action-skip">{action}</span>'
    elif "跳批检验" in action:
        action_html = f'<span class="action-skip-inspect">{action}</span>'
    else:
        action_html = f'<span class="action-normal">{action}</span>'

    st.divider()
    st.markdown("##### 当前记录预览")

    header_cols = ["到货日期","订单号","供应商","零件料号","生产日期","来料总数量","检验数量",
                   "累计批次数","执行动作","开始时间","结束时间","检验用时(分钟)","结果","检验员"]
    header_en = ["Arrival","Order No.","Supplier","Part number","Prod. Date","Total Qty","Insp. Qty",
                 "Cum. Lots","Action","Start","End","Time(min)","Result","Inspector"]
    weights = [0.95,1.1,0.85,1.4,0.95,0.8,0.8,0.75,1.3,0.65,0.65,0.85,0.6,0.85]

    hcols = st.columns(weights)
    for i, hc in enumerate(hcols):
        hc.markdown(f'<div class="tbl-header">{header_cols[i]}<br>'
                    f'<span style="font-size:10px;font-weight:normal;">{header_en[i]}</span></div>',
                    unsafe_allow_html=True)

    pn_disp = part_number if part_number else "—"
    values = [str(arrival_date), order_no if order_no else "—", supplier, pn_disp, str(production_date),
              str(int(total_qty)), str(int(inspect_qty)), str(cumulative),
              action_html, start_time.strftime("%H:%M"), end_time.strftime("%H:%M"),
              f"{diff_min:.0f}", result, inspector if inspector else "—"]
    dcols = st.columns(weights)
    for i, dc in enumerate(dcols):
        dc.markdown(f'<div class="tbl-cell">{values[i]}</div>', unsafe_allow_html=True)

    if not action:
        # 非白名单供应商：不显示跳批/正常的提示框
        pass
    elif "跳批（不检验尺寸）" in action:
        st.success("✅ 本批跳过尺寸检验，只做外观 + 包装数量")
    elif "跳批检验" in action:
        st.info("🔍 跳批序列的检验批，需做全项目检验（外观+尺寸+包装数量）")
    else:
        st.warning("📋 正常检验：外观 + 尺寸 + 包装数量全检")

    st.divider()
    if st.button("💾 保存记录到 Google Sheets", type="primary", use_container_width=True):
        if not gsheet_ok:
            st.error("Google Sheets 未连接，无法保存。")
        elif not part_number:
            st.error("请选择有效料号。")
        elif not po_number.strip():
            st.error("请填写订单号。")
        elif inspector_sel=="其他" and not inspector.strip():
            st.error("请填写检验员姓名。")
        elif result == "NG" and not defect_note.strip():
            st.error("结果为 NG，请填写不良内容后再保存。")
        else:
            row = {"到货日期":str(arrival_date),"订单号":order_no,"供应商":supplier,
                   "零件料号":part_number,"零件名称":part_name,"生产日期":str(production_date),
                   "来料总数量":int(total_qty),"检验数量":int(inspect_qty),
                   "累计批次数":cumulative,"执行动作":action,
                   "开始时间":start_time.strftime("%H:%M"),"结束时间":end_time.strftime("%H:%M"),
                   "检验用时(分钟)":f"{diff_min:.0f}","结果":result,"不良备注":defect_note.strip(),
                   "检验员":inspector,"记录时间":datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            # 测量字段（未显示的留空）
            for mf in MEASURE_COLS:
                row[mf] = measure_values.get(mf, "")
            try:
                append_record(ws, row)
                # 验证：读取当前Sheet总行数
                try:
                    total = max(0, len(ws.get_all_values()) - 1)
                except Exception:
                    total = "?"
                st.success(f"✅ 记录已保存！Google Sheets 当前共 {total} 条记录。")
                st.cache_data.clear()
                st.rerun()
            except Exception as e:
                st.error(f"保存失败：{e}")

with tab2:
    st.subheader("历史检验记录")
    st.caption(f"📊 共从 Google Sheets 读取到 **{len(history_df)}** 条记录")
    if history_df.empty:
        st.info("暂无历史记录。")
    else:
        # ---------- 筛选 ----------
        colF1, colF2, colF3 = st.columns(3)
        with colF1:
            f_supplier = st.multiselect("按供应商筛选", SUPPLIERS)
        with colF2:
            all_parts = history_df["零件料号"].astype(str).unique().tolist()
            f_part = st.multiselect("按料号筛选", all_parts)
        with colF3:
            f_result = st.multiselect("按结果筛选", RESULTS)

        view = history_df.copy().reset_index(drop=True)
        # 记录原始行号，便于删除时定位
        view["_行号"] = view.index
        if f_supplier: view = view[view["供应商"].isin(f_supplier)]
        if f_part: view = view[view["零件料号"].astype(str).isin(f_part)]
        if f_result: view = view[view["结果"].isin(f_result)]

        show_cols = ["到货日期","订单号","供应商","零件料号","生产日期","来料总数量","检验数量",
                     "累计批次数","执行动作","开始时间","结束时间","检验用时(分钟)","结果","不良备注"] \
                    + MEASURE_COLS + ["检验员"]
        show_cols = [c for c in show_cols if c in view.columns]

        st.markdown("##### 📝 编辑 / 删除记录")
        st.caption("双击单元格可直接修改内容；勾选「删除」列后点下方按钮删除。修改和删除后需点「保存修改到 Google Sheets」生效。")

        # 全选删除
        select_all = st.checkbox("全选（勾选后将删除全部当前显示的记录）")

        # 构造编辑表：加一个"删除"勾选列
        editor_df = view[show_cols].copy()
        editor_df.insert(0, "删除", select_all)

        edited = st.data_editor(
            editor_df,
            use_container_width=True,
            height=420,
            num_rows="fixed",
            key="hist_editor",
            column_config={
                "删除": st.column_config.CheckboxColumn("删除", help="勾选要删除的行", width="small"),
                "结果": st.column_config.SelectboxColumn("结果", options=RESULTS, width="small"),
                "不良备注": st.column_config.TextColumn("不良备注", width="medium"),
                "检验员": st.column_config.TextColumn("检验员"),
            },
            disabled=["累计批次数", "执行动作"],  # 这两列由系统计算，不允许手改
        )

        st.divider()

        # ---------- 操作按钮 ----------
        btn1, btn2 = st.columns(2)

        with btn1:
            if st.button("💾 保存修改到 Google Sheets", type="primary", use_container_width=True):
                try:
                    new_hist = history_df.copy().reset_index(drop=True)
                    view_rows = view["_行号"].tolist()
                    editable_cols = [c for c in show_cols if c not in ("累计批次数", "执行动作")]
                    for i, orig_idx in enumerate(view_rows):
                        for c in editable_cols:
                            new_hist.at[orig_idx, c] = edited.iloc[i][c]
                    new_hist = recalculate_all(new_hist)
                    overwrite_all(ws, new_hist)
                    st.success("✅ 修改已保存，并已重算累计批次！")
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"保存修改失败：{e}")

        with btn2:
            if st.button("🗑️ 删除勾选的记录", use_container_width=True):
                try:
                    to_delete = [view["_行号"].tolist()[i]
                                 for i in range(len(edited)) if edited.iloc[i]["删除"]]
                    if not to_delete:
                        st.warning("未勾选任何记录。")
                    else:
                        new_hist = history_df.copy().reset_index(drop=True)
                        new_hist = new_hist.drop(index=to_delete).reset_index(drop=True)
                        new_hist = recalculate_all(new_hist)
                        overwrite_all(ws, new_hist)
                        st.success(f"✅ 已删除 {len(to_delete)} 条记录，并已重算累计批次！")
                        st.cache_data.clear()
                        st.rerun()
                except Exception as e:
                    st.error(f"删除失败：{e}")

        st.divider()

        # ---------- 统计 ----------
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("总记录数", len(view))
        if "执行动作" in view.columns and len(view):
            skip_cnt = view["执行动作"].astype(str).str.contains("跳批（不检验尺寸）").sum()
            c2.metric("跳过尺寸检验批数", int(skip_cnt))
        if "结果" in view.columns and len(view):
            c3.metric("NG批数", int((view["结果"]=="NG").sum()))
        if "检验用时(分钟)" in view.columns and len(view):
            try:
                total_min = pd.to_numeric(view["检验用时(分钟)"], errors="coerce").sum()
                c4.metric("累计检验用时(小时)", f"{total_min/60:.1f}")
            except Exception:
                pass

        # ---------- 导出 ----------
        exp1, exp2 = st.columns(2)
        with exp1:
            csv = view[show_cols].to_csv(index=False).encode("utf-8-sig")
            st.download_button("📥 导出 CSV", csv, "inspection_records.csv",
                               "text/csv", use_container_width=True)
        with exp2:
            try:
                xlsx_bytes = make_excel(view[show_cols])
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    "📊 导出 Excel",
                    xlsx_bytes,
                    f"来料检验记录_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"生成Excel失败：{e}")
